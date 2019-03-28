import scrapy
import openpyxl
from datetime import datetime
import os

class PSXSpider(scrapy.Spider):
    name = "psx"

    def start_requests(self):
        current_date = datetime.today()
        input_rows = []

        path = '../data/{:04d}_{:02d}_{:02d}.xlsx'.format(current_date.year, current_date.month, current_date.day)

        if os.path.isfile(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'symbol'

        urls = []

        with open('urls.txt', 'r') as urls_file:
            for line in urls_file:
                urls.append(line.strip())

        col_num = 0
        timestamp = '{:02d}:{:02d}'.format(current_date.hour, current_date.minute)
        if str(ws.cell(1,ws.max_column).value).strip() == timestamp:
            col_num = ws.max_column
        elif str(ws.cell(1,ws.max_column).value).strip() == '':
            col_num = ws.max_column
            ws.cell(1,col_num).value = timestamp
        else:
            col_num = ws.max_column + 1
            ws.cell(1,col_num).value = timestamp

        for url in urls:
            request = scrapy.Request(url=url, callback=self.parse)
            request.meta['workbook'] = wb
            request.meta['col_num'] = col_num
            request.meta['path'] = path
            yield request

    def parse(self, response):
        wb = response.meta['workbook']
        col_num = response.meta['col_num']
        path = response.meta['path']
        ws = wb.active

        symbol = response.css('div.pageHeader__title::text').extract_first()
        stock_value = response.css('div.quote__close::text').extract_first()[3:]

        row_num = 0
        for i in range(1, ws.max_row + 1):
            row_symbol = str(ws.cell(i,1).value).strip()
            if row_symbol == symbol:
                row_num = i
                ws.cell(i,col_num).value = stock_value
                wb.save(path)
                break
        
        if row_num == 0:
            row_num = ws.max_row + 1
            ws.cell(row_num, 1).value = symbol
            ws.cell(row_num, col_num).value = stock_value
            wb.save(path)

                