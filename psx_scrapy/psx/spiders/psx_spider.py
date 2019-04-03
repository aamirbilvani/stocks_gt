import scrapy
import openpyxl
from datetime import datetime
import os
import csv

class PSXSpider(scrapy.Spider):
    name = "psx"

    def start_requests(self):
        current_date = datetime.today()

        # weekday: 0 = Monday
        # only update between 9.30 and 15.30 on weekdays
        if current_date.weekday() < 5 and (current_date.hour == 9 and current_date.minute >= 30) or \
            current_date.hour >= 10 and current_date.hour <= 14 or \
            (current_date.hour == 15 and current_date.minute <= 35):

            path = '../data/{:04d}_{:02d}_{:02d}.xlsx'.format(current_date.year, current_date.month, current_date.day)
            # xtra_path = '../data/symbols.csv'

            if os.path.isfile(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = 'symbol'

            urls = []

            with open('urls.txt', 'r') as urls_file:
                for line in urls_file:
                    urls.append(line.strip())

            row_num = 0
            timestamp = '{:02d}:{:02d}'.format(current_date.hour, current_date.minute)
            if str(ws.cell(ws.max_row, 1).value).strip() == timestamp:
                row_num = ws.max_row
            elif str(ws.cell(ws.max_row, 1).value).strip() == '':
                row_num = ws.max_row
                ws.cell(row_num, 1).value = timestamp
            else:
                row_num = ws.max_row + 1
                ws.cell(row_num, 1).value = timestamp

            for url in urls:
                request = scrapy.Request(url=url, callback=self.parse)
                request.meta['workbook'] = wb
                request.meta['row_num'] = row_num
                request.meta['path'] = path
                # request.meta['xtra_path'] = xtra_path
                yield request

    def parse(self, response):
        wb = response.meta['workbook']
        row_num = response.meta['row_num']
        path = response.meta['path']
        xtra_path = response.meta['xtra_path']
        ws = wb.active

        # with open(xtra_path, 'a') as symbols:
        #     symbols_csv = csv.writer(symbols)

        symbol = response.css('div.pageHeader__title::text').extract_first()
        stock_value = response.css('div.quote__close::text').extract_first()[3:]

        # symbols_row = [symbol, response.url]
        # symbols_row.append(response.css('div.quote__name::text').extract_first())
        # symbols_csv.writerow(symbols_row)

        col_num = 0
        for i in range(1, ws.max_column + 1):
            col_symbol = str(ws.cell(1, i).value).strip()
            if col_symbol == symbol:
                col_num = i
                ws.cell(row_num, i).value = float(stock_value.replace(',',''))
                wb.save(path)
                break
        
        if col_num == 0:
            col_num = ws.max_column + 1
            ws.cell(1, col_num).value = symbol
            ws.cell(row_num, col_num).value = float(stock_value.replace(',',''))
            wb.save(path)

                